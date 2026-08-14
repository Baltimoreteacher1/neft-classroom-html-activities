"""Emit SY26-27 Grade 6 Math Pacing Map.xlsx"""

import json, datetime as dt, xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from plan_core import *

OUT = "/Users/joelneft/Desktop/SY26-27 Grade 6 Math Planning Calendar/SY26-27 Grade 6 Math Pacing Map.xlsx"
P = json.load(open("plan.json"))
rows = P["rows"]
for r in rows:
    r["d"] = dt.date.fromisoformat(r["date"])

HDR = PatternFill("solid", fgColor="1F3864")
HF = Font(bold=True, color="FFFFFF", size=11)
TITLE = Font(bold=True, size=13, color="1F3864")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
wb = Workbook()


def sheet(name, headers, data, widths, freeze="A3", title=None, table_name=None):
    ws = wb.create_sheet(name)
    ws["A1"] = title or name
    ws["A1"].font = TITLE
    ws.append([])
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=c)
        cell.fill, cell.font = HDR, HF
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30
    for row in data:
        ws.append(row)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=len(headers)):
        for c in row:
            c.border = BOX
            if c.row > 3:
                c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = freeze
    if table_name and ws.max_row > 3:
        ref = f"A3:{get_column_letter(len(headers))}{ws.max_row}"
        t = Table(displayName=table_name, ref=ref)
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws.add_table(t)
    ws.page_setup.orientation = "landscape"
    ws.print_title_rows = "3:3"
    return ws


# ------------------------------------------------------------ 1. Year Pacing
H = [
    "Date",
    "Day",
    "Week",
    "Quarter",
    "School Status",
    "Instructional Minutes Type",
    "Unit",
    "Unit Name",
    "Lesson ID",
    "Lesson Title",
    "Standard",
    "Day Type",
    "Objective",
    "Interactive Whole Group Lesson",
    "Small-Group Lessons",
    "Calendar Notes",
    "Planning Notes",
]
data = []
for r in rows:
    data.append(
        [
            r["d"],
            r["d"].strftime("%a"),
            r["week"] or None,
            r["quarter"],
            r["status"],
            r["minutes"],
            (f"Unit {r['ewl_unit']}" if r["ewl_unit"] else (r["unit_key"] or "")),
            r["unit_label"],
            r["lesson_id"],
            r["title"],
            r["standard"],
            r["day_type"],
            r["objective"],
            r["lesson_url"],
            r["sg_urls"],
            r["cal_note"],
            r["note"],
        ]
    )
ws = sheet(
    "Year Pacing",
    H,
    data,
    [11, 6, 7, 8, 22, 12, 8, 30, 9, 40, 10, 15, 44, 34, 44, 34, 34],
    freeze="C4",
    title="Year Pacing — one row per school date (Aug 24, 2026 – Jun 11, 2027)",
    table_name="YearPacing",
)
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=1):
    row[0].number_format = "yyyy-mm-dd"
# subtle status shading (status text is always present too — never colour-only)
SH = {
    "Holiday / School Closed": "F2F2F2",
    "Break": "EDEDED",
    "PD — No Students": "F2F2F2",
    "Wellness Day": "F2F2F2",
    "Half Day / Early Release": "FFF6E5",
}
for i, r in enumerate(rows, start=4):
    f = SH.get(r["status"])
    if f:
        for c in range(1, len(H) + 1):
            ws.cell(row=i, column=c).fill = PatternFill("solid", fgColor=f)

# ------------------------------------------------------- 2. Scope & Sequence
b = xlrd.open_workbook("/Users/joelneft/Downloads/Course 1 scope and sequence.xls")
s = b.sheet_by_name("Course 1")
ss = []
for rr in range(1, s.nrows):
    v = [s.cell_value(rr, c) for c in range(s.ncols)]
    wk = v[0]
    date = v[1]
    if isinstance(date, float) and date > 1000:
        date = dt.date(1899, 12, 30) + dt.timedelta(days=int(date))
    ss.append(
        [
            str(wk).replace(".0", "") if not isinstance(wk, str) else wk,
            date,
            v[2] if v[2] != "" else None,
            str(v[3]).replace("\n", " "),
            v[4] or None,
            v[5] or None,
            str(v[6]),
            str(v[7]),
            str(v[8]),
            str(v[9]),
        ]
    )
ss = ss[1:]
wsS = sheet(
    "Scope and Sequence",
    [
        "Week",
        "Boundary Date",
        "Instructional Days",
        "Unit Guidance Document",
        "Instructional Components",
        "Additional Unit Days",
        "Curriculum Assessments",
        "District Assessments",
        "MSTAR Practice Tasks",
        "Notes and Updates",
    ],
    ss,
    [10, 14, 14, 42, 14, 13, 20, 18, 20, 30],
    title="Scope & Sequence — verbatim structured copy of 'Course 1' (original .xls untouched)",
    table_name="ScopeSeq",
)
for row in wsS.iter_rows(min_row=4, max_row=wsS.max_row, min_col=2, max_col=2):
    row[0].number_format = "yyyy-mm-dd"

# ------------------------------------------------------------ 3. Unit Summary
us, first = [], 4
for u in P["units"]:
    key = u["key"]
    ur = [r for r in rows if r["unit_key"] == key]
    n = len(ur)
    # Unit blocks are contiguous in DATE order but contain interleaved no-school
    # rows, so derive the sheet range from real row indices, not a running count.
    idx = [i for i, r in enumerate(rows) if r["unit_key"] == key]
    lo, hi = idx[0] + 4, idx[-1] + 4
    C = lambda t: (
        f"=COUNTIFS('Year Pacing'!$G$4:$G${len(rows) + 3},$A{len(us) + 4},'Year Pacing'!$L$4:$L${len(rows) + 3},\"{t}\")"
    )
    us.append(
        [
            (f"Unit {u['ewl']}" if u["ewl"] else key),
            u["label"],
            u["ewl"],
            dt.date.fromisoformat(u["start"]),
            dt.date.fromisoformat(u["end"]),
            u["budget"],
            n,
            f"=COUNTIF('Year Pacing'!$L${lo}:$L${hi},\"Core Lesson\")",
            f"=COUNTIF('Year Pacing'!$L${lo}:$L${hi},\"Core Lesson\")+COUNTIF('Year Pacing'!$L${lo}:$L${hi},\"Continued Lesson\")",
            f"=COUNTIF('Year Pacing'!$L${lo}:$L${hi},\"Assessment\")",
            f"=COUNTIF('Year Pacing'!$L${lo}:$L${hi},\"Project\")",
            f"=COUNTIF('Year Pacing'!$L${lo}:$L${hi},\"Review\")",
            f"=COUNTIF('Year Pacing'!$L${lo}:$L${hi},\"Flex\")+COUNTIF('Year Pacing'!$L${lo}:$L${hi},\"Catch-Up\")",
            u["marker"] or "—",
        ]
    )
    first += n
wsU = sheet(
    "Unit Summary",
    [
        "Unit",
        "District Unit Name",
        "EWL Unit #",
        "Start",
        "End",
        "S&S Day Budget",
        "Scheduled Dates",
        "Core Lessons",
        "Core+Continued",
        "Assessment",
        "Project",
        "Review",
        "Flex / Catch-Up",
        "S&S Assessment Marker",
    ],
    us,
    [10, 42, 11, 12, 12, 14, 14, 16, 15, 12, 10, 10, 14, 18],
    title="Unit Summary — totals are formula-driven from the Year Pacing sheet",
)
last = wsU.max_row
wsU.append(
    [
        "TOTAL",
        "",
        "",
        "",
        "",
        f"=SUM(F4:F{last})",
        f"=SUM(G4:G{last})",
        "",
        f"=SUM(I4:I{last})",
        f"=SUM(J4:J{last})",
        f"=SUM(K4:K{last})",
        f"=SUM(L4:L{last})",
        f"=SUM(M4:M{last})",
        "",
    ]
)
for c in range(1, 15):
    wsU.cell(row=last + 1, column=c).font = Font(bold=True)
for row in wsU.iter_rows(min_row=4, max_row=wsU.max_row, min_col=4, max_col=5):
    for c in row:
        c.number_format = "yyyy-mm-dd"

# --------------------------------------------------------- 4. Calendar Events
ev = [
    [
        dt.date(2026, 8, 18),
        dt.date(2026, 8, 21),
        "PD — No Students",
        PRE_YEAR_PD[2],
        "Before first student day",
    ]
]
for d in sorted(CLOSURES):
    ev.append([d, d, CLOSURES[d][0], CLOSURES[d][1], ""])
for d in sorted(HALF_DAYS):
    ev.append(
        [d, d, "Half Day / Early Release", HALF_DAYS[d], "Shortened instructional day"]
    )
for q, a, bb in QUARTERS:
    ev.append(
        [
            a,
            bb,
            "Quarter",
            f"{q} instructional span",
            "Quarter start is source-derived; end = day before next quarter start",
        ]
    )
for d in sorted(MILESTONES):
    ev.append([d, d, "Milestone", MILESTONES[d], ""])
for a, bb in CONFERENCE_WINDOWS:
    ev.append([a, bb, "Conference Window", "Parent-teacher conference window", ""])
ev.append(
    [
        TESTING_WINDOW[0],
        TESTING_WINDOW[1],
        "Testing Window",
        "State testing window — calendar labels it 'MCAP Window'; the scope & sequence calls the test MSTAR",
        "Exact in-building testing dates are NOT in either source — confirm locally",
    ]
)
ev.sort(key=lambda x: (x[0], x[2]))
wsE = sheet(
    "Calendar Events",
    ["Start", "End", "Event Type", "Description", "Planning Impact"],
    ev,
    [12, 12, 24, 62, 52],
    title="Calendar Events — every non-standard school day, extracted from the SY26-27 calendar",
    table_name="CalEvents",
)
for row in wsE.iter_rows(min_row=4, max_row=wsE.max_row, min_col=1, max_col=2):
    for c in row:
        c.number_format = "yyyy-mm-dd"

# -------------------------------------------------------------- 5. Flex Days
fx = []
for r in rows:
    if r["day_type"] in ("Flex", "Catch-Up", "Review"):
        why = r["note"] or ""
        if r["minutes"] == "Half Day":
            why += " · Placed on a shortened day by design"
        if r["day_type"] == "Review":
            why = why or "Pre-assessment retrieval"
        fx.append(
            [
                r["d"],
                r["week"],
                r["quarter"],
                f"Unit {r['ewl_unit']}" if r["ewl_unit"] else r["unit_key"],
                r["day_type"],
                r["title"],
                why.strip(" ·"),
                r["lesson_url"],
            ]
        )
wsF = sheet(
    "Flex Days",
    ["Date", "Week", "Quarter", "Unit", "Type", "Title", "Purpose", "Resource"],
    fx,
    [12, 7, 8, 10, 12, 34, 52, 40],
    title="Flex Days — every deliberate buffer, catch-up and review day in the year",
    table_name="FlexDays",
)
for row in wsF.iter_rows(min_row=4, max_row=wsF.max_row, min_col=1, max_col=1):
    row[0].number_format = "yyyy-mm-dd"

# -------------------------------------------------------------- 6. Validation
N = len(rows) + 3
va = [[n, st, det] for n, st, det in P["validation"]]
va.append(
    ["— Live workbook checks below recalculate whenever you edit Year Pacing —", "", ""]
)
live = [
    (
        "Instructional dates (Full + Half)",
        f"=COUNTIF('Year Pacing'!$E$4:$E${N},\"Full Instructional Day\")+COUNTIF('Year Pacing'!$E$4:$E${N},\"Half Day / Early Release\")",
    ),
    (
        "Instructional day-value (half days count 0.5)",
        f"=COUNTIF('Year Pacing'!$E$4:$E${N},\"Full Instructional Day\")+0.5*COUNTIF('Year Pacing'!$E$4:$E${N},\"Half Day / Early Release\")",
    ),
    (
        "Core lesson days scheduled",
        f"=COUNTIF('Year Pacing'!$L$4:$L${N},\"Core Lesson\")",
    ),
    (
        "Continued (multi-day) lesson days",
        f"=COUNTIF('Year Pacing'!$L$4:$L${N},\"Continued Lesson\")",
    ),
    ("Assessment days", f"=COUNTIF('Year Pacing'!$L$4:$L${N},\"Assessment\")"),
    ("Project days", f"=COUNTIF('Year Pacing'!$L$4:$L${N},\"Project\")"),
    ("Review days", f"=COUNTIF('Year Pacing'!$L$4:$L${N},\"Review\")"),
    (
        "Flex + Catch-Up days",
        f"=COUNTIF('Year Pacing'!$L$4:$L${N},\"Flex\")+COUNTIF('Year Pacing'!$L$4:$L${N},\"Catch-Up\")",
    ),
    ("MSTAR / testing days", f"=COUNTIF('Year Pacing'!$L$4:$L${N},\"MCAP / Testing\")"),
    ("No-instruction days", f"=COUNTIF('Year Pacing'!$L$4:$L${N},\"No Instruction\")"),
    (
        "EXCEPTION — lessons placed on a no-school day (must be 0)",
        f'=SUMPRODUCT((\'Year Pacing\'!$L$4:$L${N}<>"No Instruction")*(ISNUMBER(MATCH(\'Year Pacing\'!$E$4:$E${N},{{"Holiday / School Closed";"Break";"PD — No Students";"Wellness Day"}},0))))',
    ),
    (
        "EXCEPTION — new-content launch on a shortened day (should be 0)",
        f"=COUNTIFS('Year Pacing'!$F$4:$F${N},\"Half Day\",'Year Pacing'!$L$4:$L${N},\"Core Lesson\")",
    ),
    (
        "EXCEPTION — instructional day with no assignment (must be 0)",
        f"=COUNTIFS('Year Pacing'!$F$4:$F${N},\"<>No Students\",'Year Pacing'!$J$4:$J${N},\"\")",
    ),
    (
        "Remaining uncommitted capacity (Flex days only)",
        f"=COUNTIF('Year Pacing'!$L$4:$L${N},\"Flex\")",
    ),
]
wsV = sheet(
    "Validation",
    ["Check", "Result", "Detail"],
    va,
    [64, 12, 60],
    title="Validation — build-time checks (static) and live workbook checks (formulas)",
)
wsV.append([])
r0 = wsV.max_row + 1
wsV.cell(row=r0, column=1, value="LIVE COUNTS AND EXCEPTIONS").font = Font(
    bold=True, color="1F3864"
)
for i, (name, f) in enumerate(live, start=r0 + 1):
    wsV.cell(row=i, column=1, value=name).border = BOX
    c = wsV.cell(row=i, column=2, value=f)
    c.border = BOX
    if name.startswith("EXCEPTION"):
        wsV.cell(row=i, column=1).font = Font(bold=True)
        wsV.cell(
            row=i,
            column=3,
            value="Non-zero means the plan has drifted — fix before use",
        ).border = BOX
for i in range(4, wsV.max_row + 1):
    v = wsV.cell(row=i, column=2).value
    if v == "FAIL":
        wsV.cell(row=i, column=2).font = Font(bold=True, color="9C0006")

del wb["Sheet"]
wb.save(OUT)
print("wrote", OUT)
print("sheets:", wb.sheetnames)
